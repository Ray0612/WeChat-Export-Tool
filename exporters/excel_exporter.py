# -*- coding: utf-8 -*-
"""Excel exporter"""
import datetime

def export(msgs, path, my_name="我", title=None, session_title=None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    wb=Workbook(); ws=wb.active; ws.title="聊天记录"
    ws.column_dimensions['A'].width=20; ws.column_dimensions['B'].width=18
    ws.column_dimensions['C'].width=60; ws.column_dimensions['D'].width=10
    hf=Font(bold=True,color="FFFFFF",size=11); hfill=PatternFill(start_color="4F46E5",end_color="4F46E5",fill_type="solid")
    ha=Alignment(horizontal='center',vertical='center')
    bd=Border(left=Side(style='thin',color='E5E7EB'),right=Side(style='thin',color='E5E7EB'),top=Side(style='thin',color='E5E7EB'),bottom=Side(style='thin',color='E5E7EB'))
    ws.append(['时间','发送者','消息内容','类型'])
    for c in ws[1]: c.font=hf; c.fill=hfill; c.alignment=ha; c.border=bd
    for m in msgs:
        ts=m.get('create_time','');c=m.get('message_content','')or''
        sr=m.get('sender_username','');im=m.get('is_mine',0)
        lt=int(m.get('local_type',0))
        if lt not in(1,244813135921)or not c.strip():continue
        t=datetime.datetime.fromtimestamp(int(ts)).strftime('%Y-%m-%d %H:%M:%S')if ts.isdigit()else ts
        dn='我'if im else sr
        ws.append([t,dn,c,'文字'if lt==1 else'ZSTD'])
        r=ws.max_row
        for ci in range(1,5):
            cell=ws.cell(row=r,column=ci)
            cell.alignment=Alignment(vertical='top',wrap_text=(ci==3))
            cell.border=bd
    ws.freeze_panes='A2'; ws.auto_filter.ref='A1:D'+str(ws.max_row)
    wb.save(path)
    return path
